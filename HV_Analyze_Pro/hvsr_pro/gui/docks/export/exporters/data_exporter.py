"""
Data Exporter
=============

Functions for exporting HVSR results to CSV and JSON formats.
"""

import csv
import json
from datetime import datetime
from typing import Any, Optional, Tuple
import numpy as np
from scipy import interpolate as scipy_interpolate


def interpolate_curve(
    orig_frequencies: np.ndarray,
    curve: Optional[np.ndarray],
    n_points: int
) -> Tuple[np.ndarray, Optional[np.ndarray]]:
    """
    Interpolate curve to new frequency points.
    
    Args:
        orig_frequencies: Original frequency array
        curve: Curve to interpolate (can be None)
        n_points: Number of output points
        
    Returns:
        Tuple of (new_frequencies, interpolated_curve)
    """
    # Create new frequency array (log-spaced for HVSR)
    new_frequencies = np.logspace(
        np.log10(orig_frequencies[0]),
        np.log10(orig_frequencies[-1]),
        n_points
    )
    
    if curve is None or len(curve) == 0:
        return new_frequencies, np.full(n_points, np.nan)
    
    f = scipy_interpolate.interp1d(
        orig_frequencies, curve, kind='linear',
        bounds_error=False, fill_value='extrapolate'
    )
    return new_frequencies, f(new_frequencies)


def export_csv(filename: str, result: Any, options: dict) -> None:
    """
    Export HVSR results to CSV file.
    
    Args:
        filename: Output file path
        result: HVSRResult object
        options: Export options (may contain 'n_points' for interpolation)
    """
    # Get original data - use explicit None checks to avoid numpy array truth value issues
    orig_frequencies = result.frequencies if hasattr(result, 'frequencies') else getattr(result, 'frequency', None)
    
    mean_curve = getattr(result, 'mean_hvsr', None)
    if mean_curve is None:
        mean_curve = getattr(result, 'mean_curve', None)
    
    std_curve = getattr(result, 'std_hvsr', None)
    if std_curve is None:
        std_curve = getattr(result, 'std_curve', None)
    
    median_curve = getattr(result, 'median_hvsr', None)
    if median_curve is None:
        median_curve = getattr(result, 'median_curve', None)
    
    # Check if interpolation is needed
    n_points = options.get('n_points')
    if n_points and n_points != len(orig_frequencies):
        frequencies, mean_curve = interpolate_curve(orig_frequencies, mean_curve, n_points)
        _, std_curve = interpolate_curve(orig_frequencies, std_curve, n_points)
        if median_curve is not None:
            _, median_curve = interpolate_curve(orig_frequencies, median_curve, n_points)
    else:
        frequencies = orig_frequencies
    
    # Write CSV
    with open(filename, 'w', newline='') as f:
        writer = csv.writer(f)
        
        # Header
        header = ['Frequency (Hz)', 'Mean H/V', 'Std H/V']
        if median_curve is not None:
            header.append('Median H/V')
        writer.writerow(header)
        
        # Data rows
        for i, freq in enumerate(frequencies):
            row = [freq]
            row.append(mean_curve[i] if mean_curve is not None else '')
            row.append(std_curve[i] if std_curve is not None else '')
            if median_curve is not None:
                row.append(median_curve[i])
            writer.writerow(row)


def export_json(filename: str, result: Any, windows: Any, options: dict) -> None:
    """
    Export HVSR results to JSON file.
    
    Args:
        filename: Output file path
        result: HVSRResult object
        windows: WindowCollection object (optional)
        options: Export options (may contain 'n_points' for interpolation)
    """
    # Get original data - use explicit None checks
    orig_frequencies = result.frequencies if hasattr(result, 'frequencies') else getattr(result, 'frequency', None)
    
    mean_curve = getattr(result, 'mean_hvsr', None)
    if mean_curve is None:
        mean_curve = getattr(result, 'mean_curve', None)
    
    std_curve = getattr(result, 'std_hvsr', None)
    if std_curve is None:
        std_curve = getattr(result, 'std_curve', None)
    
    median_curve = getattr(result, 'median_hvsr', None)
    if median_curve is None:
        median_curve = getattr(result, 'median_curve', None)
    
    # Check if interpolation is needed
    n_points = options.get('n_points')
    if n_points and n_points != len(orig_frequencies):
        frequencies, mean_curve = interpolate_curve(orig_frequencies, mean_curve, n_points)
        _, std_curve = interpolate_curve(orig_frequencies, std_curve, n_points)
        if median_curve is not None:
            _, median_curve = interpolate_curve(orig_frequencies, median_curve, n_points)
    else:
        frequencies = orig_frequencies
    
    # Build JSON data
    data = {
        'metadata': {
            'export_date': datetime.now().isoformat(),
            'n_points': len(frequencies),
        },
        'frequencies': frequencies.tolist() if hasattr(frequencies, 'tolist') else list(frequencies),
        'mean_hvsr': mean_curve.tolist() if mean_curve is not None and hasattr(mean_curve, 'tolist') else (list(mean_curve) if mean_curve is not None else None),
        'std_hvsr': std_curve.tolist() if std_curve is not None and hasattr(std_curve, 'tolist') else (list(std_curve) if std_curve is not None else None),
    }
    
    if median_curve is not None:
        data['median_hvsr'] = median_curve.tolist() if hasattr(median_curve, 'tolist') else list(median_curve)
    
    # Add peak info if available
    if hasattr(result, 'primary_peak') and result.primary_peak:
        data['primary_peak'] = {
            'frequency': float(result.primary_peak.frequency),
            'amplitude': float(result.primary_peak.amplitude)
        }
    elif hasattr(result, 'peak_frequency') and result.peak_frequency:
        data['primary_peak'] = {
            'frequency': float(result.peak_frequency),
            'amplitude': float(getattr(result, 'peak_amplitude', 0))
        }
    
    # Add window statistics if available
    if windows:
        data['window_stats'] = {
            'total_windows': windows.n_windows,
            'active_windows': windows.n_active,
            'rejected_windows': windows.n_rejected,
            'acceptance_rate': windows.acceptance_rate
        }
    
    # Write JSON
    with open(filename, 'w') as f:
        json.dump(data, f, indent=2)


def export_excel(filename: str, result: Any, windows: Any, options: dict) -> None:
    """
    Export HVSR results to Excel (.xlsx) file with multiple sheets.
    
    Sheets:
        HVSR Curve: frequency, mean, median, std, percentile_16, percentile_84
        Peaks: frequency, amplitude, type, prominence
        Metadata: station, processing parameters, window stats
    
    Args:
        filename: Output file path (.xlsx)
        result: HVSRResult object
        windows: WindowCollection object (optional)
        options: Export options (may contain 'n_points' for interpolation)
    """
    try:
        import openpyxl
    except ImportError:
        raise ImportError(
            "openpyxl is required for Excel export. "
            "Install with: pip install openpyxl"
        )
    
    wb = openpyxl.Workbook()
    
    # --- Sheet 1: HVSR Curve ---
    ws_curve = wb.active
    ws_curve.title = "HVSR Curve"
    
    # Get data
    orig_frequencies = result.frequencies if hasattr(result, 'frequencies') else getattr(result, 'frequency', None)
    mean_curve = getattr(result, 'mean_hvsr', getattr(result, 'mean_curve', None))
    std_curve = getattr(result, 'std_hvsr', getattr(result, 'std_curve', None))
    median_curve = getattr(result, 'median_hvsr', getattr(result, 'median_curve', None))
    p16 = getattr(result, 'percentile_16', None)
    p84 = getattr(result, 'percentile_84', None)
    
    # Interpolate if requested
    n_points = options.get('n_points')
    if n_points and n_points != len(orig_frequencies):
        frequencies, mean_curve = interpolate_curve(orig_frequencies, mean_curve, n_points)
        _, std_curve = interpolate_curve(orig_frequencies, std_curve, n_points)
        if median_curve is not None:
            _, median_curve = interpolate_curve(orig_frequencies, median_curve, n_points)
        if p16 is not None:
            _, p16 = interpolate_curve(orig_frequencies, p16, n_points)
        if p84 is not None:
            _, p84 = interpolate_curve(orig_frequencies, p84, n_points)
    else:
        frequencies = orig_frequencies
    
    # Header
    header = ['Frequency (Hz)', 'Mean H/V', 'Std H/V']
    if median_curve is not None:
        header.append('Median H/V')
    if p16 is not None:
        header.append('Percentile 16')
    if p84 is not None:
        header.append('Percentile 84')
    ws_curve.append(header)
    
    # Data
    for i, freq in enumerate(frequencies):
        row = [float(freq)]
        row.append(float(mean_curve[i]) if mean_curve is not None else None)
        row.append(float(std_curve[i]) if std_curve is not None else None)
        if median_curve is not None:
            row.append(float(median_curve[i]))
        if p16 is not None:
            row.append(float(p16[i]))
        if p84 is not None:
            row.append(float(p84[i]))
        ws_curve.append(row)
    
    # --- Sheet 2: Peaks ---
    ws_peaks = wb.create_sheet("Peaks")
    ws_peaks.append(['Frequency (Hz)', 'Amplitude', 'Type', 'Prominence', 'Width'])
    
    if hasattr(result, 'peaks') and result.peaks:
        for peak in result.peaks:
            ws_peaks.append([
                float(peak.frequency),
                float(peak.amplitude),
                getattr(peak, 'peak_type', 'F0'),
                float(getattr(peak, 'prominence', 0)),
                float(getattr(peak, 'width', 0)),
            ])
    
    # --- Sheet 3: Metadata ---
    ws_meta = wb.create_sheet("Metadata")
    ws_meta.append(['Parameter', 'Value'])
    
    ws_meta.append(['Export Date', datetime.now().isoformat()])
    ws_meta.append(['Frequency Points', len(frequencies)])
    ws_meta.append(['Valid Windows', getattr(result, 'valid_windows', '')])
    ws_meta.append(['Total Windows', getattr(result, 'total_windows', '')])
    
    if windows:
        ws_meta.append(['Active Windows', windows.n_active])
        ws_meta.append(['Rejected Windows', windows.n_rejected])
        ws_meta.append(['Acceptance Rate', f"{windows.acceptance_rate:.1%}"])
    
    # Processing params
    params = getattr(result, 'processing_params', {})
    if params:
        ws_meta.append([])
        ws_meta.append(['Processing Parameters', ''])
        for key, value in params.items():
            ws_meta.append([key, str(value)])
    
    # Save
    wb.save(filename)
